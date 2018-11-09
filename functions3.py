import os
import os.path
import re
from ase import io
from collections import defaultdict
from pymatgen.io.vasp.outputs import Vasprun
from pymatgen.core.structure import Element
from pymatgen.core.structure import Structure
from pymatgen.symmetry.analyzer import SpacegroupAnalyzer
import operator
from operator import itemgetter
from mpinterfaces.utils import add_vacuum
from pymatgen.io.vasp.inputs import Incar
from decimal import Decimal
from pymatgen.io.vasp.inputs import Poscar
from pymatgen.io.vasp.outputs import Procar
from pymatgen.io.vasp.inputs import Potcar
from pymatgen.ext.matproj import MPRester
from pymatgen.io.vasp.inputs import Kpoints
from openpyxl import Workbook
from openpyxl.chart import (ScatterChart,Reference,Series)
import numpy as np
import subprocess
from pymatgen.transformations.standard_transformations import RotationTransformation
from mpinterfaces.utils import get_magmom_string
from mpinterfaces.mat2d.electronic_structure.startup import *
from collections import Counter
from pymatgen.core.composition import Composition

pseudos = {'H':'H','He':'He','Li':'Li_sv','Be':'Be','B':'B','C':'C','N':'N',
'O':'O','F':'F','Ne':'Ne','Na':'Na_pv','Mg':'Mg','Al':'Al','Si':'Si','P':'P','S':'S'
,'Cl':'Cl','Ar':'Ar','K':'K_sv','Ca':'Ca_sv','Sc':'Sc_sv','Ti':'Ti_sv','V':'V_sv'
,'Cr':'Cr_pv','Mn':'Mn_pv','Fe':'Fe','Co':'Co','Ni':'Ni','Cu':'Cu','Zn':'Zn','Ga':'Ga_d','Ge':'Ge_d',
'As':'As','Se':'Se','Br':'Br','Kr':'Kr','Rb':'Rb_sv','Sr':'Sr_sv','Y':'Y_sv','Zr':'Zr_sv','Nb':'Nb_sv','Mo':'Mo_sv',
'Tc':'Tc_pv','Ru':'Ru_pv','Rh':'Rh_pv','Pd':'Pd','Ag':'Ag','Cd':'Cd','In':'In_d','Sn':'Sn_d','Sb':'Sb','Te':'Te','I':'I',
'Xe':'Xe','Cs':'Cs_sv','Ba':'Ba_sv', 'La':'La','Ce':'Ce','Pr':'Pr_3','Nd':'Nd_3','Pm':'Pm_3','Sm':'Sm_3','Eu':'Eu_2','Gd':'Gd_3','Tb':'Tb_3',
'Dy':'Dy_3','Ho':'Ho_3','Er':'Er_3','Tm':'Tm_3','Yb':'Yb_2','Lu':'Lu_3','Hf':'Hf_pv','Ta':'Ta_pv','W':'W_pv','Re':'Re'
,'Os':'Os','Ir':'Ir','Pt':'Pt','Au':'Au','Hg':'Hg','Tl':'Tl_d','Pb':'Pb_d','Bi':'Bi_d','Po':'Po_d','At':'At_d','Rn':'Rn','Fr':'Fr_sv'
,'Ra':'Ra_sv','Ac':'Ac','Th':'Th','Pa':'Pa','U':'U','Np':'Np','Pu':'Pu','Am':'Am','Cm':'Cm'}
Elem_radius = {i:Element(i).atomic_radius for i in pseudos.keys()}

ncl_submit_file = """#!/bin/bash
#SBATCH --job-name=Mo2Cl6_2D_spin_x
#SBATCH --output=job_%j.out
#SBATCH --mail-type=fail
#SBATCH --mail-user=djordje.gluhovic@ufl.edu
#SBATCH --account=hennig
#SBATCH --qos=hennig
#SBATCH --nodes=1
#SBATCH --ntasks-per-node=16
#SBATCH --ntasks-per-socket=8
#SBATCH --distribution=cyclic:cyclic
#SBATCH --mem-per-cpu=2000mb
#SBATCH --time=02:00:00
module load intel/2016.0.109
module load openmpi/1.10.2
module load fftw/3.3.4
module load vasp/5.4.1
VASP=vasp_ncl
srun --mpi=pmi2 $VASP  > job.log 2>&1"""

submit_file_noz = """#!/bin/bash
#PBS -N mp-28117_2d_magnetism
#SBATCH -o job_%j.out   #Name output file
#SBATCH --mail-type=fail   #What emails you want
#SBATCH --mail-user=djordje.gluhovic@ufl.edu   #Where
#SBATCH --qos=hennig
#SBATCH --ntasks=32                  # Number of MPI ranks
#SBATCH --cpus-per-task=1            # Number of cores per MPI rank
#SBATCH --nodes=2                    # Number of nodes
#SBATCH --ntasks-per-node=16         # How many tasks on each node
#SBATCH --ntasks-per-socket=8        # How many tasks on each CPU or socket
#SBATCH --distribution=cyclic:cyclic
#SBATCH --partition=hpg1-compute
#SBATCH --mem-per-cpu=2000mb   #Per processor memory request
#SBATCH -t 10:00:00   #Walltime in hh:mm:ss or d-hh:mm:ss


module purge
module load intel/2017.1.132
module load openmpi/1.10.2
module load fftw/3.3.4
srun --mpi=pmi2 /home/mashton/vasp.5.4.4/bin/vasp_noz  > job.log"""

submit_file = """#!/bin/bash
#PBS -N mp-28117_2d_magnetism
#SBATCH -o job_%j.out   #Name output file
#SBATCH --mail-type=fail   #What emails you want
#SBATCH --mail-user=djordje.gluhovic@ufl.edu   #Where
#SBATCH --qos=hennig
#SBATCH --ntasks=32                  # Number of MPI ranks
#SBATCH --cpus-per-task=1            # Number of cores per MPI rank
#SBATCH --nodes=2                    # Number of nodes
#SBATCH --ntasks-per-node=16         # How many tasks on each node
#SBATCH --ntasks-per-socket=8        # How many tasks on each CPU or socket
#SBATCH --distribution=cyclic:cyclic
#SBATCH --partition=hpg1-compute
#SBATCH --mem-per-cpu=2000mb   #Per processor memory request
#SBATCH -t 10:00:00   #Walltime in hh:mm:ss or d-hh:mm:ss


module purge
module load intel/2017.1.132
module load openmpi/1.10.2
module load fftw/3.3.4
srun --mpi=pmi2 /home/mashton/vasp.5.4.4/bin/vasp  > job.log"""
                                           
def get_name():
    if os.path.exists('POSCAR'):
        POS = open('POSCAR','r').readlines()
        name = POS[5].split()
        num = POS[6].split()
        if int(num[1]) > int(num[0]):
            Name = name[0]+num[0]+name[1]+num[1]
        else:
            Name = name[1]+num[1]+name[0]+num[0]
        return Name

def formation_energy(twoD_DIR,threeD_DIR):
    twoList = []
    threeList = []
    Titel = 'MPs 2D_energy 3D_energy formation_energy\n'
    form = open('formation_energies','w')
    form.write(Titel)
    
    os.chdir(twoD_DIR)
    two_info = open('energy_info','r').readlines()
    for i in range(len(two_info)):
        twoList.append(two_info[i].split())
        
    os.chdir(threeD_DIR)
    three_info = open('energy_info','r').readlines()        
    for i in range(len(three_info)):
        threeList.append(three_info[i].split()) 
               
    for i in range(len(twoList)):
        for j in range(len(threeList)):
            if twoList[i][0] == threeList[j][0]:
                two_energy = float(twoList[i][1])
                three_energy = float(threeList[i][1])
                two_num = float(num(os.path.join(twoD_DIR,twoList[i][0])))
                three_num = float(num(os.path.join(threeD_DIR,threeList[i][0])))
                form_energy = (two_energy/two_num-three_energy/three_num)*1000
                form.write(twoList[i][0]+'  '+twoList[i][1]+'   '+threeList[j][1]+'  '+str(form_energy)+'\n')
    form.close()
    
        
def check_band():
    if os.path.exists('OSZICAR'):
        OSZ = open('OSZICAR','r').readlines()
        if OSZ and  'E0= ' in OSZ[len(OSZ)-1]:
            return 1
        else: return 0
    else: return 0

def num(DIR):
    A = os.getcwd()
    os.chdir(DIR)
    if os.path.exists('POSCAR'):
        sume= 0
        line = open('POSCAR','r').readlines()
        numb = re.findall('-?[0-9]+', line[6])
        for i in range(len(numb)):
            sume +=int(numb[i])
        os.chdir(A)
        return sume
    else:
        os.chdir(A)
        raise ValueError('NO POSCAR found')
        


    
def check_if_done():
    if os.path.exists('OUTCAR'):
            OUT = open('OUTCAR','r').readlines()
            l =len(OUT)
            for i in xrange(l-1,int(90*l/100),-1):
                    if 'reached required accuracy' in OUT[i]:
                                return True
            return 0
    return 0
    
def proper():
    if os.path.exists('OSZICAR'):
        OSZ = open('OSZICAR','r').readlines()
        if len(OSZ) < 3:
            print ('OSZICAR empty')
            return 0
        numb = re.findall('^ *[0-9]+', OSZ[len(OSZ)-1])
        number_of_ionic_steps = int(re.sub(' *','',numb[0]))
        if number_of_ionic_steps > 5:
            return 0
        else: return 1
    else:
        return 0
        
def check_CONTCAR():
    if os.path.exists('CONTCAR'):
        CO = open('CONTCAR','r').read()
        if CO:
            os.system('mv CONTCAR POSCAR')
            return 1
        else: return 0
    else: return 0
    
def extract_energy():
    OSZ = open('OSZICAR','r').readlines()
    sake = re.sub('.*E0= ','',OSZ[len(OSZ)-1])
    sake = re.sub(' *d .*\n','',sake)
    sake = float(sake)
    return sake

            
def symmetrize():
    structure = Structure.from_file('POSCAR')
    symm = SpacegroupAnalyzer(structure).get_refined_structure()
    prim = SpacegroupAnalyzer(symm).find_primitive()
    prim.to(fmt="poscar", filename="POSCAR_symm")
    
def nice_format(File,Titel_present = False):
    if os.path.exists(File):
        content = open(File,'r').readlines()
        if not Titel_present:
            new_file = open(File,'w')
            for i in range(len(content)):
                string = ''
                temp = content[i].split()
                for i in range(len(temp)):
                    if len(temp[i]) >12:
                         string +=temp[i]+(20 - len(temp[i]))*' '
                    else:
                        string +=temp[i]+(15 - len(temp[i]))*' '
                string +='\n'
                new_file.write(string)
            new_file.close()
        
        else:
            titel = content[0]
            new_file = open(File,'w')
            new_file.write(titel)
            for i in range(1,len(content)):
                string = ''
                temp = content[i].split()
                for i in range(len(temp)):
                    if len(temp[i]) >12:
                         string +=temp[i]+(20 - len(temp[i]))*' '
                    else:
                        string +=temp[i]+(15 - len(temp[i]))*' '
                string +='\n'
                new_file.write(string)
            new_file.close()
    else: raise IOError(File+' does not exist')

def check_if_HSE_done():
    if os.path.exists('OUTCAR'):
            OUT = open('OUTCAR','r').readlines()
            l =len(OUT)
            for i in xrange(l-1,int(98*l/100),-1):
                    if 'General timing and accounting informations for this job:' in OUT[i]:
                                return True
            return 0
    return 0
    
def create_energy_dic_anti_fero_magnetism(DIR):
    A = os.getcwd()
    DIC = {}
    for child in os.listdir(DIR):
        path = os.path.join(DIR,child,'anti_fero_magnetism')
        if os.path.isdir(path):
            os.chdir(path)
            if check_if_done() and proper():
                DIC.update({child: extract_energy()})
    os.chdir(A)
    return DIC

def formation_energy_anti_fero_magnetism(twoD_DIR,threeD_DIR):
    twoList = []
    threeList = []

    os.chdir(twoD_DIR)
    Titel = 'MPs 2D_energy 3D_energy formation_energy\n'
    form = open('formation_energies_anti_fero_magnetism','w')
    form.write(Titel)
    two_info = open('energy_info_anti_fero_magnetism','r').readlines()
    for i in range(len(two_info)):
        twoList.append(two_info[i].split())

    os.chdir(threeD_DIR)
    three_info = open('energy_info_anti_fero_magnetism','r').readlines()
    for i in range(len(three_info)):
        threeList.append(three_info[i].split())

    for i in range(len(twoList)):
        for j in range(len(threeList)):
            if twoList[i][0] == threeList[j][0]:
                form_energy = (float(twoList[i][1])-float(threeList[j][1]))/num(os.path.join(twoD_DIR,twoList[i][0]))*1000
                form.write(twoList[i][0]+'  '+twoList[i][1]+'   '+threeList[j][1]+'  '+str(form_energy)+'\n')
    form.close()
    
def create_energy_dic_magnetism(DIR):
    DIC = {}
    for child in os.listdir(DIR):
        path = os.path.join(DIR,child,'magnetism')
        if os.path.isdir(path):
            os.chdir(path)
            if check_if_done() and proper():
                DIC.update({child: extract_energy()})
    return DIC
    
def formation_energy_magnetism(twoD_DIR,threeD_DIR):
    twoList = []
    threeList = []

    os.chdir(twoD_DIR)
    Titel = 'MPs 2D_energy 3D_energy formation_energy\n'
    form = open('formation_energies_magnetism','w')
    form.write(Titel)
    two_info = open('energy_info_magnetism','r').readlines()
    for i in range(len(two_info)):
        twoList.append(two_info[i].split())

    os.chdir(threeD_DIR)
    three_info = open('energy_info_magnetism','r').readlines()
    for i in range(len(three_info)):
        threeList.append(three_info[i].split())

    for i in range(len(twoList)):
        for j in range(len(threeList)):
            if twoList[i][0] == threeList[j][0]:
                form_energy = (float(twoList[i][1])-float(threeList[j][1]))/num(os.path.join(twoD_DIR,twoList[i][0]))*1000
                form.write(twoList[i][0]+'  '+twoList[i][1]+'   '+threeList[j][1]+'  '+str(form_energy)+'\n')
    form.close()

def write_energy_from_dic(choice,DIC = {}):
    if choice == 0:
        raise SystemExit

    if choice ==1:
        energy = open('energy_info','w')
    if choice ==2:
        energy = open('energy_info_magnetism','w')
    if choice ==3:
        energy = open('energy_info_anti_fero_magnetism','w')
    if choice ==4:
        energy = open('energy_info_spin_x','w')
        
    if choice ==5:
        energy = open('energy_info_spin_z','w')
        
    for key,value in DIC.iteritems():
        energy.write(key+'  '+str(value)+'\n')
    energy.close()
    
def extract_magnetisation():
    if check_if_done() and proper():
        OUT = open('OUTCAR','r').readlines()
        l = len(OUT)
        for i in xrange(l-1,0,-1):
            if ' magnetization (x)' in OUT[i]:
                break
        for j in xrange(i,l-1):
            if 'tot  ' in OUT[j]:
                total_magnetization = OUT[j].split()
                break
        print ('  '.join(total_magnetization))
  
def duplicate_job_delete(ID_file):
    if os.path.exists(ID_file):
        jobs = open(ID_file,'r').readlines()
    else:
        raise IOError('No '+ID_file+' file found.')
    List = []
    redundent = []
    for k in range(len(jobs)):
        List.append(jobs[k].split())
        
    for i in range(len(List)-1):
        for j in xrange(i+1,len(List)):
            if List[i][1] == List[j][1]:
                if List[i][2] == 'R' and List[j][2] == 'Q':
                    redundent.append(List[j][0])
                    break
                if List[j][2] == 'R' and List[i][2] == 'Q':
                    redundent.append(List[i][0])
                    break
                if List[j][2] == 'R' and List[i][2] == 'R':
                    redundent.append(List[i][0])
                    redundent.append(List[j][0])
                    break
                if List[j][2] == 'Q' and List[i][2] == 'Q':
                    redundent.append(List[j][0])
                    break
    for k in range(len(redundent)):
        os.system('qdel '+redundent[k])
    
def name_substitute(substitute):
    if os.path.exists('submit'):
        SUB = open('submit','r').read()
        SUB = re.sub('#PBS -N.*\n','#PBS -N '+substitute+'\n',SUB)
        New_sub = open('submit','w')
        New_sub.write(SUB)
        New_sub.close()
    else:
        raise IOError('No submit file found')
        

    
def CONTCAR_proper_2D():
    LIST = []
    try:
        CON = open('CONTCAR','r').readlines()
        if not CON:
            print ('CONTCAR is empty')
            return False
        LIST.append(CON[2].split())
        LIST.append(CON[3].split())
        LIST.append(CON[4].split())
        for i in range(len(LIST)):
            for j in range(2):
                if abs(float(LIST[i][j])) > 10.0:
                    return False
                    
        for i in range(len(LIST)):
            if abs(float(LIST[i][2])) > 42.0:
                return False
        return True
    except:
        CUR_DIR = os.getcwd().split('/')
        print ('NO CONTCAR present in '+CUR_DIR[len(CUR_DIR)-1])
        try:
            CON = open('POSCAR','r').readlines()
            if not CON:
                print ('POSCAR is empty')
                return False
            LIST.append(CON[2].split())
            LIST.append(CON[3].split())
            LIST.append(CON[4].split())
            for i in range(len(LIST)):
                for j in range(2):
                    if abs(float(LIST[i][j])) > 10.0:
                        return False
                    
            for i in range(len(LIST)):
                if abs(float(LIST[i][2])) > 40.0:
                    return False
            return True
        except:
            CUR_DIR = os.getcwd().split('/')
            print ('NO POSCAR present in '+CUR_DIR[len(CUR_DIR)-1])
            
def CONTCAR_proper_3D():
    LIST = []
    try:
        CON = open('CONTCAR','r').readlines()
        if not CON:
            print ('CONTCAR is empty')
            return False
        LIST.append(CON[2].split())
        LIST.append(CON[3].split())
        LIST.append(CON[4].split())
        for i in range(len(LIST)):
            for j in range(2):
                if abs(float(LIST[i][j])) > 10.0:
                    return False
                    
        for i in range(len(LIST)):
            if abs(float(LIST[i][2])) > 25.0:
                return False
        return True
    except:
        CUR_DIR = os.getcwd().split('/')
        print ('NO CONTCAR present in '+CUR_DIR[len(CUR_DIR)-1])
        try:
            CON = open('POSCAR','r').readlines()
            if not CON:
                print ('POSCAR is empty')
                return False
            LIST.append(CON[2].split())
            LIST.append(CON[3].split())
            LIST.append(CON[4].split())
            for i in range(len(LIST)):
                for j in range(2):
                    if abs(float(LIST[i][j])) > 10.0:
                        return False
                    
            for i in range(len(LIST)):
                if abs(float(LIST[i][2])) > 25.0:
                    return False
            return True
        except:
            CUR_DIR = os.getcwd().split('/')
            print ('NO POSCAR present in '+CUR_DIR[len(CUR_DIR)-1])


            
def energy_comperator():
    LIST, LIST_mag, LIST_ant = [],[],[]
    comp = open('compared_energies','w')
    form = open('formation_energies','r').readlines()
    form_mag = open('formation_energies_magnetism','r').readlines()
    form_ant = open('formation_energies_anti_fero_magnetism','r').readlines()
    
    for i in range(1,len(form)):
        LIST.append(form[i].split())
        
    for i in range(1,len(form_mag)):
        LIST_mag.append(form_mag[i].split())
        
    for i in range(1,len(form_ant)):
        LIST_ant.append(form_ant[i].split())
        
    for i in range(len(LIST)):
        for j in range(len(LIST_mag)):
            for k in range(len(LIST_ant)):
                if LIST[i][0] == LIST_mag[j][0] and LIST[i][0] == LIST_ant[k][0]:
                    
                    index, value = min(enumerate([float(LIST[i][1]),float(LIST_mag[j][1]), float(LIST_ant[k][1])]), key=operator.itemgetter(1))
                    
                    if index == 0:
                        sub_str = 'relax'
                    elif index ==1:
                        sub_str = 'magne'
                    else:
                        sub_str = 'anti_'
                    
                    string = LIST[i][0] + ' ' +str(value) + ' ' + sub_str+' '
                    
                    index, value = min(enumerate([float(LIST[i][2]),float(LIST_mag[j][2]), float(LIST_ant[k][2])]), key=operator.itemgetter(1))
                    
                    if index == 0:
                        sub_str1 = 'relax'
                    elif index ==1:
                        sub_str1 = 'magne'
                    else:
                        sub_str1 = 'anti_'
                    
                    string+= str(value)+ ' ' + sub_str1 + ' '
                    
                    
                    index, value = min(enumerate([float(LIST[i][3]),float(LIST_mag[j][3]), float(LIST_ant[k][3])]), key=operator.itemgetter(1))
                    
                    if index == 0:
                        sub_str2 = 'relax'
                    elif index ==1:
                        sub_str2 = 'magne'
                    else:
                        sub_str2 = 'anti_'
                    
                    string+= str(value)+ ' ' + sub_str2 + '\n'
                    comp.write(string)
                    break
    comp.close()
    
def job_id():
    Dir = open('job_id_active.txt','w')
  #  print 'Job IDs are writen in job_id_active.txt'
    os.system('qstat -a >> jobs.txt')
    count = 0
    if os.path.exists('jobs.txt'):
        jobs= open('jobs.txt','r').readlines()
        for i in range(1,len(jobs)):
            re.sub(' +',' ',jobs[i])
            lis = jobs[i].split()
            if len(lis)<4:
                continue
            if 'djordje'in lis[1]:
                #print lis
                A = lis[0]+' '+lis[3]+'  '+lis[9]+'\n'
                Dir.write(A)
                count +=1
    Dir.close()
    os.system('rm jobs.txt')
    print ('number of jobs is %d'%count)


def check_if_active(DIR,File):
    os.chdir(DIR)
    if os.path.exists('Active_directories'):
        AC = open('Active_directories','r').read()
        if File[0:20] in AC:
            return 1
        else:
            return 0
            
def new_job_id():
    os.system('qstat -u djordje.gluhovic > ID.txt')
    c=0
    ID = open('ID.txt','r').readlines()
    text = open('job_id_active.txt','w')
    for i in range(5,len(ID)):
        lis = ID[i].split()
      #  print lis
        A = lis[0]+' '+lis[3]+'  '+lis[9]+'\n'
        text.write(A)
        c+=1
    text.close()
    os.system('rm ID.txt')
    print ('number of jobs is %d'%c)
    
def find_line(word,FILE):#Used to check if 2D and 3D have the same POTCAR, by searchig for word 'TITEL'
    LIST = []
    with open(FILE) as search:
            for line in search:
                    line = line.rstrip()  # remove '\n' at end of line
                    if word in line:
                        LIST.append(line)
    return LIST
    
def copy_DIR(origin,target):
    LIST = ['KPOINTS','INCAR','POSCAR','POTCAR','WAVECAR','OSZICAR','OUTCAR','CHGCAR']
    if os.path.isdir(origin) and os.path.isdir(target):
        for i in range(len(LIST)):
            os.system('cp '+origin+'/'+LIST[i]+' '+target+'/'+LIST[i])
    else:
        raise ValueError('Origin or targer is not a directory')
    
def add_vacuum(vacuum = 20):
    struct_2d = Structure.from_file('POSCAR')
    new_struct = add_vacuum_padding(struct_2d, vacuum)
    new_struct.to(filename='POSCAR', fmt='poscar')
    
def backup_DIR(origin,target):
    if os.path.isdir(origin) and os.path.isdir(target):
        name = origin.split('/')
        os.chdir(target)
        if not os.path.exists(name[len(name)-1]):
            os.mkdir(name[len(name)-1])
        path = os.path.join(target,name[len(name)-1])
        LIST = ['KPOINTS','INCAR','POSCAR','POTCAR','WAVECAR','OSZICAR','OUTCAR','CHGCAR']
    
        for i in range(len(LIST)):
            if not os.path.isdir(os.path.join(origin,LIST[i])):
                os.system('cp '+origin+'/'+LIST[i]+' '+path+'/'+LIST[i])
            else:
                print (LIST[i])
    else:
        raise ValueError('Origin or targer is not a directory')
        
        
        
def compare_magnetization(DIR, tollerance = 0.1):
    twoD = '2D'
    threeD = '3D'
    MAG2, MAG3 = [], []
    temp = []
    os.chdir(os.path.join(DIR,twoD))
    mag_file2 = open('magnet_info','r').readlines()
    
    os.chdir(os.path.join(DIR,threeD))
    mag_file3 = open('magnet_info','r').readlines()
    
    for i in range(1,len(mag_file2)):
        A = mag_file2[i].split()
        temp.append(A[0].split('_')[0])
        temp.append(A[5])
        MAG2.append(temp)
        temp = []
       
    for i in range(1,len(mag_file3)):
        A = mag_file3[i].split()
        temp.append(A[0].split('_')[0])
        temp.append(A[5])
        MAG3.append(temp)
        temp = []
        
    for i in range(len(MAG2)):
        for j in range(len(MAG3)):
            if MAG2[i][0] == MAG3[j][0]:
                diff = float(MAG2[i][1]) - float(MAG3[j][1])
                if abs(diff) > tollerance:
                    print ('Magnetization changed from ' +MAG3[j][1]+' to' + MAG2[i][1])
                    
                    
def delete_all_jobs(only_active = False):
    new_job_id()
    File = open('job_id_active.txt','r').readlines()
    if not only_active:
        for i in range(len(File)):
            os.system('qdel '+ File[i].split()[0])
    else:
       for i in range(len(File)):
           status = File[i].split()[2]
           if status !='R':
               os.system('qdel '+ File[i].split()[0])   
            
           
def round_numbers(File, tolerance = 0.1, columns = [], start_line = 0, end = 0):
    try:
        FILE = open(File,'r').readlines()
    except IOError:
        raise IOError('No '+File +' found')
    List = []
    temp = []
    if end == 0:
        end = len(FILE)
    for i in range(start_line,end):
        A = FILE[i].split()
        for k in range(len(A)):
            temp.append(A[k])
        for j in range(len(columns)):
            temp.append('.')
        List.append(temp)
        temp =[]
    count = len(FILE[start_line].split()) #number of columns of the file 
    for i in range(len(columns)): #checking 
        if (int(columns[i])-1) > (count):
            raise ValueError('Some of the desired rows exceed the dimension of the file')
    
    
    sa = 0
    for j in range(len(columns)):
        for i in range(len(List)):
              if abs(int(float(List[i][columns[j]])) - float(List[i][columns[j]])) <= tolerance:
                  List[i][sa+count] = str(int(float(List[i][columns[j]])))
                  
              elif abs(int(float(List[i][columns[j]])) - float(List[i][columns[j]])+1) <= tolerance:
                  List[i][sa+count] = str(int(float(List[i][columns[j]]))+1)
              elif abs(int(float(List[i][columns[j]])) - float(List[i][columns[j]])-1) <= tolerance:
                  List[i][sa+count] = str(int(float(List[i][columns[j]]))-1)
              else:
                  List[i][sa+count] = str(round(float(List[i][columns[j]]),5))
        sa+=1 
        
    print (List) 
    NEW_file = open(File,'r').read()
    sa = 0 
    for j in range(len(columns)):
        for i in range(len(List)):
            NEW_file= NEW_file.replace((List[i][columns[j]]),str((List[i][sa+count])))
        sa+=1
    rounded = open(File+'_rounded','w')
    rounded.write(NEW_file)
    rounded.close()
    nice_format(File+'_rounded')
                  
def get_NBANDS_string(DIR):
    try:
        os.chdir(DIR)
    except:
        raise ValueError('No directory for get_NBANDS_string found')
    if os.path.exists('OUTCAR'):
        OUT = open('OUTCAR','r').readlines()
        for i in range(len(OUT)):
            if 'NBANDS' in OUT[i]:
                numb = re.findall('-?[0-9]+', OUT[i])
                break
        return 2*int(numb[len(numb)-1])
    else:
        raise IOError('NO OUTCAR found')
        
def get_magmom_spin(direction,yield_magnetization):
    if not os.path.exists('OSZICAR'):
        raise IOError('No OSZICAR present.')
    OSZ= open('OSZICAR','r').readlines()
    l = len(OSZ)-1
    mage = OSZ[l]
    mage = re.sub('.+mag=\s+','',mage)
    mage = re.sub('\n','',mage)
    POS = open('POSCAR','r').readlines()
    num = POS[6].split()
    if int(num[0]) < int(num[1]):
        scale = int(num[0])
        majority = int(num[1])
    else:
        scale = int(num[1])
        majority = int(num[0])
    
    magnetism = float(mage)/scale
    
    
    string = ('0 0 '+str(magnetism)+' ')*scale + ' 0 0 0'*majority
    if yield_magnetization:
        return magnetism, string
    else:
        return string
    
        
def create_spin_polarized_comp_all_directories(DIR,direction, force_overwrite = False, submit = True):
    LIST = ['KPOINTS','POSCAR','POTCAR','CONTCAR','submit']
    if 'x' in direction:
        saxis = '1 0 0'
    else:
        saxis = '0 0 1'
    
    MAG_INCAR_DICT = {'PREC': 'Accurate', 'ENCUT': 600, 'IBRION': -1,'ISPIN': 2, 'NSW ': 0, 'ISIF': 2,'ISYM': 0,'LMAXMIX':4,
     'LCHARG': False, 'ISMEAR': 1,'LWAVE': False,'NPAR': 8, 'EDIFF': 1e-8,'SIGMA': 0.1,'LSORBIT': '.True.',
    'ICHARG':11,'SAXIS': saxis}
    
    First_step_INCAR = {'PREC':'Accurate','ENCUT':450,'IBRION':2,'NSW':0,'ISIF':3,'LCHARG':'True','ISMEAR':1,'SIGMA':0.1,'EDIFF':'1E-6'
    ,'NPAR':8,'LWAVE':'True'}
   
    for child in os.listdir(DIR):
        First_phase_done = True #initialization
        path = os.path.join(DIR,child,'magnetism')
        if os.path.isdir(path):
            os.chdir(path)
            if check_if_done() and proper():
                magnetization , temp = get_magmom_spin()
                if int(magnetization) < 0.5:
                    continue
                MAG_INCAR_DICT.update({'MAGMOM': get_magmom_spin(direction,A=False)})
                MAG_INCAR_DICT.update({'NBANDS': get_NBANDS_string(path)})
                os.chdir(os.path.join(DIR,child))
                if not os.path.exists('spin_polarized'+'_'+direction):
                    os.mkdir('spin_polarized'+'_'+direction)
                    First_phase_done = False
                spin_path = os.path.join(DIR,child,'spin_polarized'+'_'+direction)
                os.chdir(spin_path)
                #Check if the first phase is done
                if os.path.exists('INCAR'):
                    INC = open('INCAR','r').read()
                    if 'LMAXMIX' not in INC:
                        if not check_if_done():
                            First_phase_done = False
                    else:
                        if check_if_done():
                            if not force_overwrite:
                                continue
                        else:
                            print (child+' '+'spin in '+direction+' not done')
                else:
                    First_phase_done = False
                if not First_phase_done:
                    First_step_INCAR.update({'NBANDS':get_NBANDS_string(path)})
                    os.chdir(spin_path)
                    Incar.from_dict(First_step_INCAR).write_file('INCAR')
                else:
                    
                    Incar.from_dict(MAG_INCAR_DICT).write_file('INCAR')
                for i in range(len(LIST)):
                    string = 'cp '+ path+'/'+LIST[i] +' .'
                    os.system(string)
                if os.path.exists('CONTCAR'):
                    os.system('mv CONTCAR POSCAR')
                name_substitute(get_name()+'_2D_spin_'+direction)
                update_submit_file(['e','p'],['/home/mashton/vasp.5.4.1/bin/vasp_ncl','32'])
                if submit:
                    os.system('sbatch submit')
                    

def update_submit_file(flags = [], values = []):
    #flags: t = time, n = name, m = memory requested, p = number of precessors requested, q = queue, e = executable, @ = email
    if len(flags) != len(values):
        raise IOError('Sizes of 2 lists do not match')
    if os.path.exists('submit'):
        sub = open('submit','r').read()
    else:
        raise IOError('No submit file present')
    for i in range(len(flags)):
        if flags[i] == 't': #time
            sub = re.sub('#SBATCH -t.*\n','#SBATCH -t '+values[i]+'   #Walltime in hh:mm:ss or d-hh:mm:ss\n',sub)
            
        elif flags [i] == 'n': #name
           sub = re.sub('#PBS -N.*\n','#PBS -N '+values[i] +'\n',sub)
        
        elif flags [i] == 'm': #memory
           sub = re.sub('#SBATCH --mem.*\n','#SBATCH --mem-per-cpu='+str(values[i]).strip(' ') +'   #Per processor memory request\n',sub)
        
        elif flags [i] == 'p':#processors
           sub = re.sub('#SBATCH --ntasks=.*\n','#SBATCH --ntasks='+str(values[i]).strip(' ')+'\n',sub)
        
        elif flags [i] == 'q':#queu
           sub = re.sub('#SBATCH --qos=.*\n','#SBATCH --qos='+values[i].strip(' ')+'\n',sub)
        
        elif flags [i] == 'e': #executable
           sub = re.sub('mpiexec.*log','mpiexec '+values[i]+'  > job.log\n',sub)
       
        elif flags [i] == '@': #executable
           sub = re.sub('#SBATCH --mail-user=.*\n','#SBATCH --mail-user='+values[i]+'\n',sub)
        
        new_sub = open('submit','w')
        new_sub.write(sub)
        new_sub.close()
        
def remove_junk(DIR,LIST = ['AECCAR*','CHG','WAVECAR','DOSCAR','PROCAR','REPORT','PCDAT','XDATCAR','EIGENVAL','job_1*','err*']):
    string = 'rm'
    for i in range(len(LIST)):
        string+= ' '+LIST[i]
    os.chdir(DIR)
    os.system(string)
    for child in os.listdir(DIR):
        path  = os.path.join(DIR,child)
        if os.path.isdir(path):
            remove_junk(path)

        
def create_spin_polarized_comp(direction, force_overwrite = False, submit = True):
    #This function should be started in ordinary relaxation directory, which containes magnetisation directory 
    LIST = ['KPOINTS','POSCAR','POTCAR','CONTCAR','submit']
    if 'x' in direction:
        saxis = '1 0 0'
    else:
        saxis = '0 0 1'
    
    MAG_INCAR_DICT = {'PREC': 'Accurate', 'ENCUT': 600, 'IBRION': -1,'ISPIN': 2, 'NSW ': 0, 'ISIF': 2,'ISYM': 0,'LMAXMIX':4,
     'LCHARG': False, 'ISMEAR': 1,'LWAVE': False,'NPAR': 4, 'EDIFF': 1e-8,'SIGMA': 0.1,'LSORBIT': '.True.',
    'ICHARG':11,'SAXIS': saxis}
    
    First_step_INCAR = {'PREC':'Accurate','ENCUT':500,'IBRION':2,'NSW':0,'ISIF':3,'LCHARG':'True','ISMEAR':1,'SIGMA':0.1,'EDIFF':'1E-6'
    ,'NPAR':4,'LWAVE':'True'}
   
    DIR = os.getcwd() # Child directory
    First_phase_done = True #initialization
    path = os.path.join(DIR,'magnetism')
    if os.path.isdir(path):
            os.chdir(path)
            if check_if_done() and proper():
                magnetization , temp = get_magmom_spin(direction,True)
                if int(magnetization) < 2:
                    return
                
                os.chdir(DIR) #create spin_polarized_X directory
                if not os.path.exists('spin_polarized'+'_'+direction):
                    os.mkdir('spin_polarized'+'_'+direction)
                    First_phase_done = False
                spin_path = os.path.join(DIR,'spin_polarized'+'_'+direction)
                os.chdir(spin_path)
                #Check if the first phase is done
                if os.path.exists('INCAR'):
                    INC = open('INCAR','r').read()
                    if 'LMAXMIX' not in INC:
                        if not check_if_done():
                            First_phase_done = False
                    else:
                        if check_if_done():
                            if not force_overwrite:
                                return
                        else:
                            print (DIR+' '+'spin in '+direction+' not done')
                else:
                    First_phase_done = False
                if not First_phase_done:
                    First_step_INCAR.update({'NBANDS':get_NBANDS_string(path)})
                    os.chdir(spin_path)
                    Incar.from_dict(First_step_INCAR).write_file('INCAR')
                else:
                    MAG_INCAR_DICT.update({'NBANDS': get_NBANDS_string(path)})
                    MAG_INCAR_DICT.update({'MAGMOM': get_magmom_spin(direction,A=False)})
                    os.chdir(spin_path)
                    Incar.from_dict(MAG_INCAR_DICT).write_file('INCAR')
                for i in range(len(LIST)):
                    string = 'cp '+ path+'/'+LIST[i] +' .'
                    os.system(string)
                if os.path.exists('CONTCAR'):
                    os.system('mv CONTCAR POSCAR')
                name_substitute(get_name()+'_2D_spin_'+direction)
                update_submit_file(['e','p'],['/home/mashton/vasp.5.4.1/bin/vasp_ncl','32'])
                if submit:
                    print ('submit')
                    #os.system('sbatch submit')
    else:
        print ('No magnetism directory in '+DIR)
        return
        
def makeKPOINTS(twoD, MeshType,Length, desired_directory):
        #Input Variables
        MeshType = MeshType
        TwoDimensional = twoD
        l = Length
        input_POSCAR = desired_directory+'/POSCAR'
        output_KPOINTS = desired_directory+'/KPOINTS'

        POSCAR = open(input_POSCAR, 'r')
        lines = POSCAR.readlines()
        scale = Decimal.from_float(float(lines[1].split()[0]))

        #Define original lattice vectors

        a1 = lines[2].split()
        a2 = lines[3].split()
        a3 = lines[4].split()

        a11 = scale*Decimal.from_float(float(a1[0]))
        a12 = scale*Decimal.from_float(float(a1[1]))
        a13 = scale*Decimal.from_float(float(a1[2]))

        a21 = scale*Decimal.from_float(float(a2[0]))
        a22 = scale*Decimal.from_float(float(a2[1]))
        a23 = scale*Decimal.from_float(float(a2[2]))

        a31 = scale*Decimal.from_float(float(a3[0]))
        a32 = scale*Decimal.from_float(float(a3[1]))
        a33 = scale*Decimal.from_float(float(a3[2]))

        #Calculate the determinant

        det = a11*(a22*a33-a23*a32)-a12*(a21*a33-a23*a31)+a13*(a21*a32-a22*a31)
        b11 = (a22*a33-a23*a32)/det
        b12 = (a21*a33-a23*a31)/det
        b13 = (a21*a32-a22*a31)/det

        b21 = (a12*a33-a13*a32)/det
        b22 = (a11*a33-a13*a31)/det
        b23 = (a11*a32-a12*a31)/det

        b31 = (a12*a23-a13*a22)/det
        b32 = (a11*a23-a13*a21)/det
        b33 = (a11*a22-a12*a21)/det

        with open(output_KPOINTS, 'w') as file:
                kpoints_x = ''
                kpoints_y = ''
                kpoints_z = ''
                file.write('Automatic mesh\n')
                file.write( '0\n')
                kpoints_x = round(float(l*Decimal.sqrt(b11*b11+b12*b12+b13*b13)))
                kpoints_y = round(float(l*Decimal.sqrt(b21*b21+b22*b22+b23*b23)))
                kpoints_z = round(float(l*Decimal.sqrt(b31*b31+b32*b32+b33*b33)))
                if TwoDimensional == True:
                        kpoints_z =1
                file.write(MeshType + '\n')
                file.write(str(kpoints_x) + " " + str(kpoints_y) + " " + str(kpoints_z))
        print (kpoints_x, kpoints_y, kpoints_z)
        
def write_POTCAR():
    os.environ['VASP_PSP_DIR']= '/home/djordje.gluhovic/POTCAR/'
    if os.path.exists('POSCAR'):
        POS = open('POSCAR','r').readlines()
        elements = POS[5].split()
        A = []
        A.append(pseudos.get(elements[0]))
        A.append(pseudos.get(elements[1]))
        POT = Potcar(A)
        POT.write_file('POTCAR')
    else:
        raise IOError('NO POSCAR present')
        
def write_POSCAR_by_mp_id(ID):
    MPR = MPRester('AFLmqRCQTI0hTHNB')
    POS = Poscar(MPR.get_structure_by_material_id(ID))
    POS.write_file('POSCAR')
    
def update_INCAR(names = [], values = []): # Check the significance of vaw line
    if len(names) != len(values):
        raise IOError('Sizes of 2 lists do not match')
        
    if not os.path.exists('INCAR'):
            raise IOError('NO INCAR present')
    INC = Incar.from_file('INCAR')        
    for i in range(len(names)):
        INC.update({names[i]:values[i]})
    
    INC.write_file('INCAR')

def write_relaxation_INCAR():
    INCAR = {'LCHARG': False, 'IBRION': 2, 'PARAM2': 0.22, 'PARAM1': 0.1833333333, 'ISMEAR': 1, 'LWAVE': False, 
    'NPAR': 8, 'SIGMA': 0.1, 'AGGAC': 0.0, 'SYSTEM': 'Cdbr2_3d', 'ENCUT': 500,
    'ISIF': 3, 'GGA': 'Bo', 'EDIFF': 1e-06, 'NSW': 60, 'LAECHG': True, 'LUSE_VDW': True, 'PREC': 'Fast'}
    INC = Incar.from_dict(INCAR)
    INC.write_file('INCAR')
    
    
def check_POTCAR():
    try:
        os.system('grep TITEL POTCAR > A')
        os.system('grep POTCAR: OUTCAR > B')
        
        A = open('A','r').read()
        B = open('B','r').readlines()
        A = re.sub(' +',' ',A)
        for i in range(len(B)):
            B[i] = re.sub('POTCAR: +','',B[i])
            B[i] = re.sub(' +',' ',B[i])
            B[i] = re.sub('\n','',B[i])
            B[i] = B[i].strip()
        if B[0] in A and B[1] in A:
            os.system('rm A B')
            return True
        else:
            os.system('rm A B')
            return False
    except:
        raise IOError('Something is wrong is '+ os.getcwd())
        
def write_KPOINTS(density):
    Str = Structure.from_file('POSCAR')
    KPO = Kpoints.automatic_density(Str, density)
    KPO.write_file('KPOINTS')
 
 
def magnetic_energy(parent_dir,magnet = 'magnetism',anti = 'anti_fero_magnetism'):
    #Returns the lower value of the two (magnetic or antifero_magnetic) from parent_directory, and the tag saying which energy is lower, the number of atoms
    
    LIST = []
    if not os.path.isdir(parent_dir):
        print ('Wrong directory submited to magnetic_formation_energy()')
        LIST.append(False)
        return LIST
   
    mag_path = os.path.join(parent_dir,magnet) 
    if not os.path.isdir(mag_path):
        print ('Parent directory is not a parent directory')
        LIST.append(False)
        return LIST
        
    ant_path = os.path.join(parent_dir,anti)
    if not os.path.isdir(ant_path):
        print ('Parent directory is not a parent directory')
        LIST.append(False)
        return LIST
    
    os.chdir(mag_path)
    if not check_if_done():
        LIST.append(False)
        return LIST
    Energy_mag = float(extract_energy()) 
    
    NUM = num(mag_path)
    
    os.chdir(ant_path)
    if not check_if_done():
        LIST.append(False)
        return LIST
    Energy_ant = float(extract_energy()) 
    
    Min = min(Energy_mag,Energy_ant)
    if Min == Energy_ant:
        tag = 'Antiferromagnetic'
    elif Min == Energy_mag:
        tag = 'Ferromagnetic'
    else:
        raise ValueError('Something is wrong with magnetic_energy()')
    LIST.append(Min)
    LIST.append(tag)
    LIST.append(NUM)
    return LIST
    
def magnetic_formation_energy(twoD_dir,threeD_dir):
    #Returns a string : 'formation energy   2D_tag    3D_tag'. Tags are describing which energy is lower (feromagnetic or antifero_magnetic)
    if not os.path.isdir(twoD_dir) or not os.path.isdir(threeD_dir):
        raise IOError('Directories provided are incorrect')

    
    monolayer_list = magnetic_energy(twoD_dir)
    bulk_list =  magnetic_energy(threeD_dir)
    
    if monolayer_list[0] == False or bulk_list[0] == False:
        return 'Wrong data present'
    
    mono_energy = monolayer_list[0]
    mono_tag = monolayer_list[1]
    mono_num = monolayer_list[2]
    
    bulk_en = bulk_list[0]
    bulk_tag = bulk_list[1]
    bulk_num = bulk_list[2]
    
    formation_energy = (float(mono_energy)/mono_num - float(bulk_en)/bulk_num)*1000
    
    return str(formation_energy)+'  '+mono_tag+ '   '+bulk_tag+'\n'
    
    
def density_of_states_to_excel():
    if not os.path.exists('DOSCAR'):
        print ('DOSCAR not present in '+os.getcwd() )
        return
        
    efermi = Vasprun('vasprun.xml').efermi
    DOS = open('DOSCAR','r').readlines()
    energy, up, down, plot_values = [],[],[],[]
    
    try:
        nedos = Incar.from_file('INCAR').as_dict()['NEDOS'] - 1
    except:
        nedos = 300
    print (nedos)
        
    try:
        for i in range(7,7+nedos+2):
            A = DOS[i].split()
            energy.append(float(float(A[0])-efermi))
            up.append(float(A[1]))
            down.append(-float(A[2]))

    except:
       print ('Error in indexing in function density_of_states_to_excel()')
    SUM = list(np.array(up)-np.array(down))

    wb = Workbook()
    ws = wb.active  # grab the active worksheet
    ws['A1'] = 'Energy' # Data can be assigned directly to cells
    ws['B1'] = 'Up'
    ws['C1'] = 'Down'
    ws['D1'] = 'Sum'

    for i in range(nedos+1):
        en_str = 'A'+str(i+2)
        up_str = 'B'+str(i+2)
        down_str = 'C'+str(i+2)
        sum_str = 'D'+str(i+2)
        ws[en_str]=float(energy[i])
        ws[up_str]=float(up[i])
        ws[down_str] = float(down[i])
        ws[sum_str] = float(SUM[i])
    chart = ScatterChart()
    chart.title = "Density of states vs normalized Fermi energy\n"+name
    chart.style = 2
    chart.legend.position = 'r'
    xvalues = Reference(ws, min_col=1, min_row=2, max_row=nedos+1)
    for i in range(2, 5):
        values = Reference(ws, min_col=i, min_row=1, max_row=nedos+1)
        series = Series(values, xvalues, title_from_data=True)
        chart.series.append(series)
    plot = wb.create_chartsheet()
    plot.add_chart(chart)

    wb.save('density_of_states.xlsx')
    
def energy_diff_magnetism(DIR):
#This function takes DIR (directory where magnetism and antiferomagnetism direccotories are located) as argument.
#It returns a list that conrains the difference in energies (feromagnetic vs antiferomagnetic) and tag saying which energy is lower 
    A = os.getcwd()
    os.chdir(DIR)
    if not os.path.exists('magnetism') or not os.path.exists('anti_fero_magnetism'):
        return ['False']
    
    magnet_path = os.path.join(DIR,'magnetism')
    anti_path = os.path.join(DIR,'anti_fero_magnetism')
    
    os.chdir(magnet_path)
    if check_if_done():
        mag_energy = extract_energy()
    else:
        mag_energy = 0
        print ('This data is not valid')
        
    
    os.chdir(anti_path)
    if check_if_done():
        anti_energy = extract_energy()
    else:
        anti_energy = 0
        print ('This data is not valid')
    difference = abs(mag_energy-anti_energy)
    if mag_energy <anti_energy:
        tag = 'magnet'
    else:
        tag = 'anti_fero_magnet'
    if mag_energy ==0 or anti_energy ==0:
        tag = False
    os.chdir(A)
    return [difference*1000,tag]
    
def extract_magnet_per_atom():
    if not os.path.exists('OSZICAR') or (not os.path.exists('POSCAR')):
        print ('OSZICAR not present in '+os.getcwd())
        return False
    OSZ = open('OSZICAR','r').readlines()
    POS = open('POSCAR','r').readlines()
    POS = map(int,POS[6].split())
    
    if POS[0] < POS[1]:
        number = POS[0]
    else:
        number = POS[1]
    
    A = OSZ[len(OSZ)-1]
    if 'F= ' in OSZ[len(OSZ)-1]:
        A=re.sub('.*mag=','',A)
        A = re.sub(' +','',A)
        return float(A)/number
    else:
        return False
        
        
def sort_against_periodic_table(DIR):
#This fucntion takes a directory, which containes all magnetic computations. It creates a list[MP-ID, name of compound, magnetization, structure].
#Name of compouns is split into elements, and elements are converted to their respective atomic numbers(M first, X second)
#List is then sorted with respect to the atomic number of metal and atomic number of halogen
#Atomic numbers are converted back to elements and elements are then combined into compounds

#Fucntion returns a sorted list[MP-ID,name of compounds, magnetization,structure]
    
    start = os.getcwd()
    if not os.path.isdir(DIR):
        raise IOError('Wrong direcctory supplied')    
    
    LIST,Sorted,temp, Final = [],[],[], []
    
    for child in os.listdir(DIR):
        path = os.path.join(DIR,child,'magnetism')
        if os.path.isdir(path):
           os.chdir(path)
           A = [child,get_name(),extract_magnet_per_atom(),SpacegroupAnalyzer(Structure.from_file('POSCAR')).get_crystal_system()]
           temp.append(A)
           
    for i in range(len(temp)):#Separation of compund into elements
        B = re.split('\d+',temp[i][1])[0:2]
        LIST.append([B[0],B[1],temp[i][0],temp[i][2],temp[i][3]])
    
    
    for element in LIST: #Conversion to Z
        element[0] = Element(element[0]).Z
        element[1] = Element(element[1]).Z
        
    Sorted = sorted(LIST, key=itemgetter(0,1)) #Sorting with respect to the first and second element of the list
    
    
    for e in Sorted:# Conversion from Z back to elements
        e[0] = Element.from_Z(e[0]).symbol
        e[1] = Element.from_Z(e[1]).symbol
        
    for e in Sorted:# Combining elements into compounds
        Final.append([e[2],e[0]+e[1]+'3',e[3],e[4]])
    os.chdir(start)
    
    return Final 
    
def sort_table_against_periodic_table(temp = []):
#Input  = list[MP-ID, name of compound, magnetization, structure] 
    LIST, Sorted, Final = [],[],[]
    for i in range(len(temp)):#Separation of compund into elements
        B = re.split('\d+',temp[i][1])[0:2]
        LIST.append([B[0],B[1],temp[i][0],temp[i][2],temp[i][3]])
    
    
    for element in LIST: #Conversion to Z
        element[0] = Element(element[0]).Z
        element[1] = Element(element[1]).Z
        
    Sorted = sorted(LIST, key=itemgetter(0,1)) #Sorting with respect to the first and second element of the list
    
    
    for e in Sorted:# Conversion from Z back to elements
        e[0] = Element.from_Z(e[0]).symbol
        e[1] = Element.from_Z(e[1]).symbol
        
    for e in Sorted:# Combining elements into compounds
        Final.append([e[2],e[0]+e[1]+'3',e[3],e[4]])
    
    return Final
    
def remove_van_der_waals():
    if not os.path.exists('INCAR'):
        raise IOError('NO Incar file found in '+os.getcwd())
    IN = Incar.from_file('INCAR').as_dict()
    try:
        IN.pop('AGGAC')
        IN.pop('PARAM2')
        IN.pop('PARAM1')
        IN.pop('@module')
        IN.pop('@class')
        IN.pop('LUSE_VDW')
        IN.pop('GGA')
    except:
        print ('Van der Waals is already partially or completely removed')    
    I = open('INCAR','w')
    for key in IN:
        I.write(str(key)+' = '+str(IN[key])+'\n')
    I.close()
    
    In = open('INCAR','r').read()
    A = In.replace('[','')
    A1 = A.replace(',','')
    B = A1.replace(']','')
    
    I = open('INCAR','w')
    I.write(B)
    I.close()
    
def update_INCAR_for_SA(): # update INCAR for spin-aniosotrpy compuations. Should be used with remove_van_der_waals() before any SA is sent to be computed
    current_DIR = os.getcwd()
    update_INCAR(['NSW','LCHARG','ENCUT','EDIFF'],['0','True','600','1E-8'])
    NBANDS_string = ''
    try:
        os.chdir('../magnetism')
        NBANDS_string = get_NBANDS_string(os.getcwd())
    except:
        print ('Wrong NBANDS string')
    os.chdir(current_DIR)
    update_INCAR(['NBANDS'],[NBANDS_string])
    
def read_PROCAR(tolerance = 0.4, ion_resolved = True):
    if not os.path.exists('PROCAR'):
       raise IOError('No PROCAR found') 
    if not os.path.exists('POSCAR'):
       raise IOError('No POSCAR found') 
       
    P = Procar('PROCAR')
    elements = open('POSCAR','r').readlines()[5].split()
    numbers = map(int,open('POSCAR','r').readlines()[6].split())
    d_orbitals = [4, 5, 6, 7, 8]
    if Element(elements[0]).is_transition_metal:
        start = 0
        end = numbers[0]
    else:
        start = numbers[0]
        end = numbers[0]+numbers[1]
    Up, Down = [],[]
    for kpoint in range(P.nkpoints):
        for band in range(P.nbands):
            if ion_resolved:
                for ion in range(start,end):
                    sume = 0
                    for orbital in d_orbitals:
                        sume = sume+P.data.items()[0][1][kpoint][band][ion][orbital]
                    if sume > tolerance:
                        Down.append([kpoint+1,band+1,ion+1,sume])
                    sume=0
                    for orbital in d_orbitals:
                        sume = sume+P.data.items()[1][1][kpoint][band][ion][orbital]
                    if sume > tolerance:
                        Up.append([kpoint+1,band+1,ion+1,sume])
            else:
                sume = 0
                for ion in range(start,end):   
                    for orbital in d_orbitals:
                        sume = sume+P.data.items()[0][1][kpoint][band][ion][orbital]
                if sume > tolerance:
                    Down.append([kpoint+1,band+1,sume])
                sume = 0
                for ion in range(start,end):    
                    for orbital in d_orbitals:
                        sume = sume+P.data.items()[1][1][kpoint][band][ion][orbital]
                if sume > tolerance:
                    Up.append([kpoint+1,band+1,sume])
    return Up,Down
    
def parse_Eigenval():
    #EIGENVAL is parsed in a list. List L can be accessed as:
    #    L[spin][kpoint][band] = [energy,occupancy]
    #Spin can be 0 (down) and 1 (up)
    if not os.path.exists('EIGENVAL'):
        raise IOError('No EIGENVAL found')
    
    Eig = open('EIGENVAL','r').readlines()
    nkpoints = int(Eig[5].split()[1])
    nbands = int(Eig[5].split()[2])
    
    Spin = [None]*2
    Kpoints_up = [None]*nkpoints
    Nbands_up = [None]*nbands # Energy and occupation per spin channel 
    
    Kpoints_down = [None]*nkpoints
    Nbands_down = [None]*nbands # Energy and occupation per spin channel 
    k_counter = 0    
    band_counter = 0
    for i in range(7,len(Eig)):
        a = re.match('^ *[1-9]+',Eig[i])
        try:
            A = a.group() #This will trigger try-except reaction if line that does not contain band information is processed. Such line will be skipped
            S = Eig[i].strip().split()
            S = map(float,S)
            Nbands_down[band_counter] = [S[2],S[4]]
            Nbands_up[band_counter] = [S[1],S[3]]
            
            band_counter = band_counter+1
            if S[0] == nbands:
                Kpoints_up[k_counter]= Nbands_up
                Kpoints_down[k_counter]= Nbands_down
                k_counter =  k_counter+1
                Nbands_down = [None]*nbands
                Nbands_up = [None]*nbands
                band_counter=0
        except:
            pass

    Spin[0] = Kpoints_down
    Spin[1] = Kpoints_up
    return [Spin[0],Spin[1]]

def read_number_of_KPOINTS():
    if not os.path.exists('PROCAR'):
        raise IOError('No PROCAR present in '+os.getcwd())
    procar_data = open('PROCAR','r').readlines()[1].split()
    number_of_kpoints = int(procar_data[3])
    return number_of_kpoints
    
def mean_eigenval_difference_spin_channels(floor = 0.2):
  #  try:
    EIG = read_EIGENVAL()
    PRO = read_PROCAR()
   # except:
     #   raise IOError('Something went wrong in '+os.getcwd())
    counter = 0
    difference = 0
    if len(EIG) != len(PRO):
        raise ValueError('PROCAR and EIGENVAL file do not have the same number of kpoints and/or bands')
    
    for i in range(len(PRO)):
        if PRO[i][1] >= floor:
            difference = difference + abs(EIG[i][1]-EIG[i][2])
            counter = counter+1
    if counter != 0:
        return difference/counter 
    else:
        return False
        
def critical_temperature(exchange_integral_cooefficinet = 6,structure = 'hexagonal'):
    # This function should be called in directory which containes magnetic and antiferromagnetic directories
    XY = {'hexagonal':0.89}
    Ising = {'hexagonal':1.519}
    temp = energy_diff_magnetism(os.getcwd())
    mae = MAE(os.getcwd())
    if (mae[0] == 'False'):
        return 0

    Kb = 8.6173303*pow(10,-5) # Boltzman constant in Ev/K
    if temp[1] == 'anti_fero_magnet':
        print ('Material in '+os.getcwd()+' has antiferromagnetic ground state and critical temperature does not make physcal sense')
        return False
    energy_difference = temp[0]/1000 # in eV
    J = energy_difference / exchange_integral_cooefficinet
    if ((mae[1].upper()) == 'X'):

        constant = XY[structure]
    else:
        constant = Ising[structure]
    critical_temperature = constant*J/Kb
    return critical_temperature
    
def get_lower_energy(extensions = ['magnetism','anti_fero_magnetism']):
    DIR = os.getcwd()
    energies = []
    for c in extensions:
        path = os.path.join(DIR,c)
        if os.path.isdir(path):
            os.chdir(path)
            try:
               energies.append(extract_energy())
            except:
                print ('Energy cannot be extracted from '+ DIR)
                os.chdir(DIR)
                return False
        else:
            os.chdir(DIR)
            print (path + 'is not a valid directory')
            return False
            
    if len(extensions) != len(energies):
        print ('Something went wrong in get_lower_energy()')
        os.chdir(DIR)
        return False
    else:
        os.chdir(DIR)
        return min(energies),extensions[energies.index(min(energies))]
        
    
        
def formation_energy_single_directory(DIR = ''):
    CUR= os.getcwd()
    if DIR != '':
        os.chdir(DIR)
    else:
        DIR = CUR
    print (DIR)
    
    if '2D' in DIR:
        base = re.sub('2D.*','',DIR)
        temp = re.findall('2D/.*$',DIR)[0]
        MP = re.sub('2D/','',temp)
    elif '3D' in DIR:
        base = re.sub('3D.*','',DIR)
        temp = re.findall('3D/.*$',DIR)[0]
        MP = re.sub('3D/','',temp)
    else:
        print ('wrong directory was supplied')
        return False
        
    extensions = ['magnetism','anti_fero_magnetism']
    dirr = ['2D','3D']
    energy_2D = 0
    energy_3D = 0
    num_2D = 0
    num_3D = 0
    LIST = []
    for d in dirr:
        LIST = []
        for e in extensions:
            path = os.path.join(base,d,MP,e)
            if os.path.isdir(path):
                os.chdir(path)
                try:
                    LIST.append(extract_energy())
                except:
                    print (path+'containse invalid OUTCAR and energy could not be extracted')
                    return False
            else:
                print (path+' does not exist')
                return False
        if d == '2D':
            energy_2D = min(LIST)
            num_2D = num(os.getcwd())
        else:
            energy_3D = min(LIST)
            num_3D = num(os.getcwd())
            
    os.chdir(CUR)
    return (energy_2D/num_2D - energy_3D/num_3D)*1000 # energy is meV
    
    
def energy_barrier_between_spin_channels(tolerance = 0.000001):
    if not os.path.exists('EIGENVAL'):
        print ('No EIGENVAL present in '+ os.getcwd())
        return False
    cbm_up = float("inf")    
    cbm_down = float("inf")
    
    V = Vasprun('vasprun.xml')
    for spin, d in V.eigenvalues.items():
        for k, val in enumerate(d):
           # occupation = val[1]
           # eigenval = val[0]
                if int(spin[0]) == -1:  #spin down 
                    if val[1] <= tolerance and val[0] < cbm_down:
                        cbm_down = val[0]    
                else:                       #spin up
                    if val[1] <= tolerance and val[0] < cbm_up:
                        cbm_up = val[0]                     
    print (cbm_up,cbm_down)
    return abs(cbm_up-cbm_down)
    
def MAE(DIR):
#This function takes DIR (directory where spin_polarized_x and spin_polarized_z direccotories are located) as argument.
#It returns a list that conrains the difference in energies (meV) (spin_polarized_x and spin_polarized_z) and tag saying which energy is lower 
    A = os.getcwd()
    os.chdir(DIR)
    if not os.path.exists('spin_polarized_x') or not os.path.exists('spin_polarized_z'):
        return ['False']
    
    x_path = os.path.join(DIR,'spin_polarized_x')
    z_path = os.path.join(DIR,'spin_polarized_z')
    
    os.chdir(x_path)
    if check_if_HSE_done():
        x_energy = extract_energy()
    else:
        print ('This data is not valid')
        os.chdir(A)
        return ['False']  
    
    os.chdir(z_path)
    if check_if_HSE_done():
        z_energy = extract_energy()
    else:
        print ('This data is not valid')
        os.chdir(A)
        return ['False']  
         
    difference = abs(x_energy-z_energy)
    if x_energy < z_energy:
        tag = 'X'
    else:
        tag = 'Z'
    os.chdir(A)
    return [difference*1000,tag]
    
def extract_number_of_ionic_steps():
    if not os.path.exists('OSZICAR'):
        raise IOError('No OSZICAR found in '+os.getcwd())
    OSZ = open('OSZICAR','r').readlines()
    L = len(OSZ)
    line = '0 0'
    for i in range(L-1,0,-1):
        if 'F=' in OSZ[i]:
            line = OSZ[i]
            break
    if line == '0 0':
        return 0
    num = line.split()
    return int(num[0])
    
def extract_time_of_job():
    if not os.path.exists('OUTCAR'):
        raise IOError('No OUTCAR found in '+os.getcwd())
        
    os.system('grep LOOP+ OUTCAR > temp')
    temp = open('temp','r').readlines()
    L = len(temp[0].split())
    total_time=0
    for i in range(len(temp)):
        time =float(temp[i].split()[L-1])
        total_time = total_time+time
    os.system('rm temp')
    return total_time
    

    
def get_POTCAR(home_dir,List = []):
    cur_dir = os.getcwd() # current directory
    counter = 1 # used to keep track to multiple POTCARs
    
    for name in List:
        string = 'find -name '+name
        os.chdir(home_dir)
        sub= subprocess.check_output(string,shell=True)
        
        if len(sub) == 0:
            os.chdir(cur_dir)
            os.system('rm -f potcar*')
            raise IOError('No POTCAR found for '+name)
            
        sub = sub.split('\n')
        for i in sub:
            if 'GGA_PAW_PBE' in i:
                path = i
        path = path.strip('./')
        path = os.path.join(home_dir,path,'POTCAR')
        os.chdir(cur_dir)
        os.system('cp '+path+' potcar'+str(counter))
        counter = counter+1
    list_of_potcars = ''
    for i in range(1,counter):
        list_of_potcars = list_of_potcars+'potcar'+str(i)+' '
    os.system('rm -f POTCAR')
    os.system('cat '+list_of_potcars+' >> POTCAR')
    os.system('rm -f potcar*')

def extract_vectors():
    if not os.path.exists('POSCAR'):
        raise IOError("No POSCAR file found in "+ os.getcwd())
    Pos = open('POSCAR','r').readlines()
    S = Structure.from_file('POSCAR')
    L = S.lattice
    gamma = L.gamma
    a,b = Pos[2],Pos[3]
    a = a.strip()
    b = b.strip()
    
    a = map(float,a.split())
    b = map(float,b.split())
    
    A_vec = np.array([a[0],a[1]])
    B_vec = np.array([b[0],b[1]])
    
    return A_vec,B_vec,gamma
    
def find(A= [],B = [], A_p=[], B_p=[]):
    minimal_diff = 0.05 # any number bigger than 0.05= 5% strain
    LIST = []
    for i in range(1,7):
        for j in range(1,7):
            for k in range(1,7):
                for l in range(1,7):
                    L1 = np.linalg.norm(i*A+j*B)
                    L2 = np.linalg.norm(k*A_p + l*B_p)
                    difference =abs((L1-L2)*2)/(L1+L2) 
                    if difference < minimal_diff:
                        LIST.append([difference,(L1+L2)/2])
    return LIST
    
def find_normal_vector(V,Z = [0,0,1]): # returns unit normal vector
    N = np.cross(np.array(V),np.array(Z))
    return N/np.linalg.norm(N)
    
def find_pareto_optimal_front(L = []): # Only works for 2 columns (x and y data)
    L = list(set(tuple(element) for element in L))
    Sorted = sorted(L, key=itemgetter(0))

    pareto_front = [Sorted[0]]
    for j in Sorted:
        if j[1] < pareto_front[len(pareto_front)-1][1]:
            pareto_front.append(j)
            
    return pareto_front
    
def get_relaxed_MAGMOM_string(DIR):
    cur_DIR = os.getcwd()
    os.chdir(DIR)
    
    a= extract_magnet_per_atom()
    POS = open('POSCAR','r').readlines()
    Name = POS[5].split()
    Num = POS[6].split()
    string = ''
    
    for i in range(len(Name)):
        if Element(Name[i]).is_transition_metal:
            string = string + str(Num[i])+ '*'+str(a)+' '
        else:
            string = string + str(Num[i])+ '*0'+' '        
    os.chdir(cur_DIR)
    return string
    
def start_spin_polarized_comp(direction, submit = True):
    files = '{"KPOINTS","POSCAR","POTCAR","CONTCAR"}' # no spaces between words and comma
        
    INCAR_DICT = {'PREC':'Accurate','ENCUT':600,'IBRION':2,'NSW':0,'ISIF':3,'LCHARG':'True','ISMEAR':1,'SIGMA':0.1,'EDIFF':'1E-8'
    ,'NPAR':4,'LWAVE':'False','ISPIN':'2','LORBIT':'11'}
   
    DIR = os.getcwd() # Parent directory
    
    path = os.path.join(DIR,'magnetism')
    if os.path.isdir(path):
            os.chdir(path)
            # If magnetic computation is done extract NBANDS and relaxed MAGMOM string
            if check_if_done():
                NBANDS= get_NBANDS_string(path)
                MAGMOM = get_relaxed_MAGMOM_string(path)
              
                os.chdir(DIR) #create spin_polarized_X directory
                if not os.path.exists('spin_polarized_'+direction):
                    os.mkdir('spin_polarized_'+direction)
                spin_path = os.path.join(DIR,'spin_polarized_'+direction)
                os.chdir(spin_path)
                
                # Copy the files (from LIST) from magnetic computation. 
                os.system('cp ../'+files+' .')
                check_CONTCAR() # If CONTCAR is non empty replace POSCAR with it
                
                #Write INCAR from INCAR_DICT
                INCAR_DICT.update({'MAGMOM':MAGMOM,'NBANDS':NBANDS})
                Incar.from_dict(INCAR_DICT).write_file('INCAR')
                
                #Update submit file
                sub = open('submit','w')
                sub.write(submit_file)
                sub.close()
                update_submit_file2({'n':get_name()+'_spin_'+direction,'t':'2:00:00'})
                
                if submit:
                    os.system('sbatch submit')
            else:
                print( 'Magnetic computation did not converge')
                return
    else:
        print ('No magnetism directory in '+DIR)
        return

def second_step_MAE(direction,submit = True): #should be run in spin_polarized_direction directory
    LIST = ['KPOINTS','POSCAR','POTCAR','CHGCAR','INCAR']
    if 'x' in direction:
        saxis = '1 0 0'
    else:
        saxis = '0 0 1'
    
    MAG_INCAR_DICT = {'PREC': 'Accurate', 'ENCUT': 600, 'IBRION': -1,'ISPIN': 2, 'NSW': 0, 'ISIF': 2,'ISYM': 0,'LMAXMIX':4,
     'LCHARG': False, 'ISMEAR': 1,'LWAVE': False,'NPAR': 4, 'EDIFF': 1e-8,'SIGMA': 0.1,'LSORBIT': '.True.',
    'ICHARG':11,'SAXIS': saxis}
    
    
    if direction =='z': # Copy necessary files from directory where MAE was started
        for l in LIST:
            os.system('cp ../spin_polarized_x/'+l+' .')
        
    if os.path.getsize('CHGCAR') < 100:
        raise ValueError('No meaningful CHGCAR found in ' +os.getcwd())
 
    os.system('cp ../magnetism/OSZICAR .')   
    MAG_INCAR_DICT.update({'MAGMOM': get_magmom_spin(direction,False)})
    update_INCAR(MAG_INCAR_DICT.keys(),MAG_INCAR_DICT.values())
    
    sub = open('submit','w')
    sub.write(ncl_submit_file)
    sub.close()
    update_submit_file(['t','n'],['2:00:00',get_name()+'_MAE_2'])
    
    if submit:
        os.system('sbatch submit')
       
def angle_between(p1, p2):
    ang1 = np.arctan2(*p1[::-1])
    ang2 = np.arctan2(*p2[::-1])
    return np.rad2deg((ang1 - ang2) % (2 * np.pi))
    

    
def diff_btw_lists(first, second): # find elemnts in fisrt list that are absent from second list. 
                                    #Usually first list  = range(x), where x is the upper boundary
    return list(set(first)- set(second))
    
class Cut_structure(Structure):
    
    def align_axis(self,axis_to_be_aligned = 'a',axis_against_which_aligment_is_done = []):
        if axis_to_be_aligned not in ['a','b']:
            raise ValueError('Non supported axis used')   
        if  not isinstance(axis_against_which_aligment_is_done,list) or not len(axis_against_which_aligment_is_done)==2: 
            raise ValueError('Wrong argument supplied for "axis_against_which_aligment_is_done"')        
        
        matrix = self.lattice.matrix
        A,B,gamma = matrix[0,0:2],matrix[1,0:2],self.lattice.gamma
        #S = Structure.from_file('POSCAR')
        x,y =  np.array([1,0]),np.array([0,1])
        
        if axis_to_be_aligned == 'a':
            axis = np.array(A)
        elif axis_to_be_aligned == 'b':
            axis = np.array(B)   
        Coor_axis = axis_against_which_aligment_is_done
        
        axis_unit_vector = axis/np.linalg.norm(axis) 
        coor_axis_unit_vector = Coor_axis/np.linalg.norm(Coor_axis)
        print (axis,Coor_axis )  
        theta1 = angle_between(axis_unit_vector,coor_axis_unit_vector)
        
    
        R = RotationTransformation([0,0,1],-theta1)
        S1 = R.apply_transformation(self)
        return S1
        
    def remove_species_below(self,direction,horizontal_offset):
        if not isinstance(direction,list) or not len(direction)==2:
            raise ValueError('Direction is a list with x and y values')
            
        if horizontal_offset > 1:
            if direction == 'a':
                horizontal_offset = horizontal_offset/self.lattice.a
            else:
                horizontal_offset = horizontal_offset/self.lattice.b
                
        if direction[0] != 0:# Accounts for the slope = infinity
            slope = float(direction[1]/direction[0])
            # d contains indicies of atoms that are below the direction vector        
            d = [count for count,i in enumerate(self.frac_coords) if i[1]< slope*(i[0]-horizontal_offset)]
        else:
            d = [count for count,i in enumerate(self.frac_coords) if i[0]> horizontal_offset]
        return self.remove_sites(d)
        
    def remove_species_above(self,direction,horizontal_offset):
        if not isinstance(direction,list) or not len(direction)==2:
            raise ValueError('Direction is a list with x and y values')
            
        if horizontal_offset > 1:
            if direction == 'a':
                horizontal_offset = horizontal_offset/self.lattice.a
            else:
                horizontal_offset = horizontal_offset/self.lattice.b
                
        if direction[0] != 0:# Accounts for the slope = infinity
            slope = float(direction[1]/direction[0])
            # d contains indicies of atoms that are below the direction vector        
            d = [count for count,i in enumerate(self.frac_coords) if i[1]> slope*(i[0]-horizontal_offset)]
        else:
            d = [count for count,i in enumerate(self.frac_coords) if i[0]< horizontal_offset]
            
        return self.remove_sites(d)
        
    def is_bonded(self,site1,site2,bond_len = 3):
        dist = np.linalg.norm(site1.coords-site2.coords)
        print (dist)
        if dist <=bond_len:
            return True
        else: return False
        
def update_INCAR_dict(DICT={}):   
    if not os.path.exists('INCAR'):
            raise IOError('NO INCAR present')
    INC = Incar.from_file('INCAR')        
    INC.update(DICT)
    INC.write_file('INCAR')

def update_submit_file2(DICT):
    #flags: t = time, n = name, m = memory requested, p = number of precessors/tasls requested, q = queue, e = executable, @ = email
    # nn = number of nodes, 
    if os.path.exists('submit'):
        sub = open('submit','r').read()
    else:
        raise IOError('No submit file present')
    for key,value in DICT.iteritems():
        if key == 't': #time
            sub = re.sub('#SBATCH -t.*\n','#SBATCH -t '+value+'   #Walltime in hh:mm:ss or d-hh:mm:ss\n',sub)
            
        elif key == 'n': #name
           sub = re.sub('#PBS -N.*\n','#PBS -N '+value +'\n',sub)
        
        elif key == 'm': #memory
           sub = re.sub('#SBATCH --mem.*\n','#SBATCH --mem-per-cpu='+str(value).strip(' ') +'   #Per processor memory request\n',sub)
        
        elif key == 'p':#processors
           sub = re.sub('#SBATCH --ntasks=.*\n','#SBATCH --ntasks='+str(value).strip(' ')+'\n',sub)
        
        elif key == 'q':#queu
           sub = re.sub('#SBATCH --qos=.*\n','#SBATCH --qos='+value.strip(' ')+'\n',sub)
        
        elif key == 'e': #executable
           sub = re.sub('mpiexec.*log','mpiexec '+value+'  > job.log\n',sub)
           sub = re.sub('srun.*log','srun --mpi=pmi2 '+value+'  > job.log\n',sub)
       
        elif key == '@': #executable
           sub = re.sub('#SBATCH --mail-user=.*\n','#SBATCH --mail-user='+value+'\n',sub)
        
        new_sub = open('submit','w')
        new_sub.write(sub)
        new_sub.close()
    
    
def create_magnetic_calculation(submit = True):
    FILES = ['POSCAR','POTCAR','INCAR','submit','KPOINTS','CONTCAR','runjob','vdw_kernel.bindat']
    if not os.path.exists('magnetism'):
        os.mkdir('magnetism')
        
    os.chdir('magnetism')
    
    os.system('cp ../{'+','.join(FILES)+'} .')
    
    if os.path.exists('CONTCAR') and os.stat('CONTCAR').st_size > 10:
        os.system('mv CONTCAR POSCAR')

    update_INCAR_dict({'LORBIT':11,'ISPIN':2,'MAGMOM':get_magmom_string(Structure.from_file('POSCAR'))})
    
    if submit:
        try:
            os.system('sbatch submit')
        except:
            os.system('sbatch runjob')
            
def create_anti_fero_magnetic_calculation(submit = True):
    FILES = ['POSCAR','POTCAR','INCAR','submit','KPOINTS','CONTCAR','runjob','vdw_kernel.bindat']
    if not os.path.exists('anti_fero_magnetism'):
        os.mkdir('anti_fero_magnetism')
        
    os.chdir('anti_fero_magnetism')
    
    os.system('cp ../{'+','.join(FILES)+'} .')
    
    if os.path.exists('CONTCAR') and os.stat('CONTCAR').st_size > 10:
        os.system('mv CONTCAR POSCAR')

    update_INCAR_dict({'LORBIT':11,'ISPIN':2,'MAGMOM':get_magmom_string_anti_fero()})
    
    if submit:
        try:
            os.system('sbatch submit')
        except:
            os.system('sbatch runjob')
            
def get_magmom_string_anti_fero():
    if not os.path.exists('POSCAR'):
        raise ValueError('No POSCAR found in '+os.getcwd())
        
    POS = open('POSCAR','r').readlines()
    Name = POS[5].strip().split()
    Num = POS[6].strip().split()
    string = ''
    if Element(Name[0]).is_transition_metal:
        a= int(Num[0])
        string += '5 -5 '*(a/2)
        string += Num[1]+'*0.5'
    else:
        string += Num[0]+'*0.5'
        a = int(Num[1])
        string += ' 5 -5'*(a/2)
    return string
    
def fast_norm(a):
#Much faster variant of numpy linalg norm
    return np.sqrt(np.dot(a, a))
    
def time():
    if check_if_done():
        sub= subprocess.check_output('grep -a "Total CPU time used (sec):" OUTCAR',shell=True)
        return float(sub.split()[-1])
    else:
        return 0

def submit_builder(file_name='submit',exe = '5.4.4',type_of_exe = 'std',update_dict = {},wannier = False):
    
    first_line = '#!/bin/bash\n'
    
    config={'--time':'10:00:00',
 '--distribution': 'cyclic:cyclic',
 '--job-name': 'Name',
 '--mail-type': 'fail',
 '--output':'job_%j.out',
 '--mail-user': 'djordjeapiss@gmail.com',
 '--mem-per-cpu': '2gb',
 '--nodes': 1,
 '--ntasks': 32,
 '--ntasks-per-core':1,
 '--ntasks-per-socket': 16,
 '--partition': 'hpg2-compute',
 '--qos': 'hennig',
 '--ntasks-per-node': 32}
    
    list_of_allowed_exe = ['5.4.4','5.4.1']
    if exe not in list_of_allowed_exe:
        print ('The version '+exe+' is not supported')
        return
    executable = re.sub('5.4.4',exe,'/home/mashton/vasp.5.4.4/bin/vasp')############################
    
    modulus = {'5.4.4':'module purge\nmodule load intel/2017.1.132\nmodule load openmpi/1.10.2\n',
    '5.4.1':'module load intel/2016.0.109 openmpi\nmodule load fftw/3.3.4\n'}

    if len(update_dict) > 0:
        config = update_sub_dict(config,update_dict)
    
    if type_of_exe == 'std':
        pass
    elif type_of_exe == 'noz':
        executable += '_noz'
    else:
        raise ValueError("No desired execuatble found")   
    
    #Writing of a file
    s = open(file_name,'w')
    s.write(first_line)
    
    for key,value in config.iteritems():
        s.write('#SBATCH '+key+'='+str(value)+'\n')
    s.write('\n'+modulus[exe])
    
    if wannier:
        s.write('module load wannier90/2.1.0\n')
    
    s.write('\nsrun --mpi=pmi2 '+executable+' > job.log')
    s.close()
    
def update_sub_dict(config,DICT):
    determinable = ['--ntasks-per-node','--nodes','--ntasks','--ntasks-per-core','--ntasks-per-socket']
    
    for key,value in DICT.iteritems():
        if key not in config.keys():
            raise IOError ('Given parameter is not supported: '+key)
    
    if '--partition' in DICT.keys(): config['--partition'] = DICT['--partition']
    
    #Processor check         

    determinable = ['--ntasks-per-node','--nodes','--ntasks','--ntasks-per-core','--ntasks-per-socket'] 
    if '--ntasks' in DICT.keys() or '--nodes' in DICT.keys() or '--ntasks-per-core' in DICT.keys() or '--ntasks-per-socket' in DICT.keys():
    
        if '--ntasks-per-node' in DICT.keys() and '--ntasks-per-socket' in DICT.keys():
            n = 2*int(DICT['--ntasks-per-socket'])
            if n != int(DICT['--ntasks-per-node']):
                raise ValueError('conflicting values')
            
        if '--ntasks' in DICT.keys() and '--nodes' in DICT.keys() and '--ntasks-per-node' in DICT.keys():
            ntask_per_nodee = int(DICT['--ntasks'])/int(DICT['--nodes'])
            if ntask_per_nodee  != int(DICT['--ntasks-per-node']):
                raise ValueError('conflicting values')
                
        if '--ntasks' in DICT.keys() and '--nodes' in DICT.keys():
            num = int(DICT['--ntasks'])%int(DICT['--nodes'])
            if num != 0 :
                raise ValueError('conflicting values')   
                
        if '--ntasks' in DICT.keys() and '--ntasks-per-node' in DICT.keys():
            numc = int(DICT['--ntasks'])%int(DICT['--ntasks-per-node'])
            if numc != 0 :
                raise ValueError('conflicting values') 
                    
        if '--ntasks' in DICT.keys() and '--ntasks' in DICT.keys() != '--ntasks' in config.keys():
            config['--ntasks'] = DICT['--ntasks']
        
        if '--nodes' in DICT.keys() and '--nodes' in DICT.keys() != '--nodes' in config.keys():
            config['--nodes'] = DICT['--nodes']
        
        if '--ntasks-per-core' in DICT.keys() and '--ntasks-per-core' in DICT.keys() != '--ntasks-per-core' in config.keys():
            config['--ntasks-per-core'] = DICT['--ntasks-per-core']
        
        if '--ntasks-per-node' not in DICT.keys():
            ntask_per_node = int(config['--ntasks'])/int(config['--nodes'])
            if (ntask_per_node>32 and config['--partition'] == 'hpg2-compute') or (ntask_per_node>64 and config['--partition'] == 'hpg1-compute'):
                raise ValueError('Requested number of tasks cannot be divided across requested number of nodes')
            config['--ntasks-per-node'] = ntask_per_node
            config['--ntasks-per-socket'] = ntask_per_node/2
            
        if '--ntasks-per-node' in DICT.keys():
            config['--ntasks-per-node'] = DICT['--ntasks-per-node']
            ntasks = int(config['--ntasks-per-node'])*int(config['--nodes'])
            config['--ntasks'] = ntasks
            config['--ntasks-per-socket'] = int(config['--ntasks-per-node'] )/2
            
        if '--ntasks-per-socket' in DICT.keys():
            ntask_per_node = int(config['--ntasks-per-socket'])*2
            config['--ntasks-per-node'] = ntask_per_node
            ntasks = int(config['--ntasks-per-node'])*int(config['--nodes'])
            config['--ntasks'] = ntasks
            
    for d in determinable:
            try:
                DICT.pop(d)
            except: pass 
     #Memory check       
    if '--mem-per-cpu' in DICT.keys():
        mem = DICT['--mem-per-cpu']
        DICT.pop('--mem-per-cpu')
    else:
        mem = config['--mem-per-cpu']
        
    mem_value = float(re.findall('[-+]?\d*\.\d+|[-+]?\d+',mem)[0])
    if 'gb' in mem: mem_value *=1024
    requested_mem = int(config['--ntasks-per-core'])*int(config['--ntasks'])*mem_value
    allowed_mem = int(config['--ntasks'])*1024*3
    if requested_mem>allowed_mem:
        raise ValueError('Requested memory cannot be provided')
    else:
        config['--mem-per-cpu'] = mem
    
    for key in DICT.keys():
        config[key] = DICT[key]
        
    return config
    

def write_wannier_input_file(ibz_path = '../'):
        
        ibz_file = os.path.join(ibz_path,'IBZKPT')
        if not os.path.isfile(ibz_file):
            raise IOError('No IBZKPT found in '+os.getcwd())
        try:
            sub= subprocess.check_output('grep Tetrahedra '+ibz_file+' -B 1',shell=True)
        except:
            sub = open(ibz_file,'r').readlines()[-1] 
         
        last_line = sub.split('\n')[0]
        os.system('mv KPOINTS kpoints_backup') # Save previous file
        write_band_structure_kpoints(Structure.from_file('POSCAR'),n_kpts=1,dim=2,ibzkpt_path = ibz_path)
        
        K = open('KPOINTS','r').readlines()
        K.reverse()
        lines = []
        
        for k in range(len(K)):
            if last_line in K[k]:
                break
            else:
                if len(K[k]) > 3:
                    lines.append(K[k])
                    
        f= defaultdict(list)           
        for l in lines:
            a = l.split(' ')
            key = a[-1].strip('\n')
            key = key.replace("\\","")
            values = a[0:3]
            f[key] = values
            
        string = """bands_plot = true
bands_plot_format gnuplot xmgrace
num_wann = 96\n"""

        IN = Incar.from_file('INCAR')
        NB = str(IN['NBANDS'])
        
        string = re.sub('96',NB,string)
        
        string += '\nbegin kpoint_path\n'
        
        LIST = []
        for key,value in f.iteritems():
            LIST.append('{} {}'.format(key,' '.join(value)))
            
        for i in range(len(LIST)-1):
            string += '{}  {}\n'.format(LIST[i],LIST[i+1])
        
        string+= '{}  {}\n'.format(LIST[i+1],LIST[0])
        string += 'end kpoint_path\n\nbands_num_points = 20\n'
        
        w = open('wannier90.win','w')
        w.write(string)
        w.close()
        
        incar = """ISMEAR = -5
EMIN = -20 ; EMAX = 20 ; NEDOS = 1000  # usefull energy range for density of states    
ALGO = None ; NELM = 1                 # exact diagonalization one step suffices
NBANDS = 96                            # need for a lot of bands in GW    
LORBIT = 11
LWANNIER90_RUN = .TRUE."""

        IN = open('INCAR','w')
        IN.write(incar)
        IN.close()
        
        os.system('mv kpoints_backup KPOINTS')

def convert_cif_to_POSCAR(DIR):
    for child in os.listdir(DIR):
        path = os.path.join(DIR,child)
        if os.path.isdir(path):
            convert_cif_to_POSCAR(path)
        if os.path.isfile(child) and 'cif' in child:
            try:
                a = io.read(child)
                a.write('POSCAR_'+child.strip('.cif'), format = 'vasp')
            except:
                print (child)

    
def remove_tags_from_INCAR(inc_file= 'INCAR',LIST = []):
    I = Incar.from_file(inc_file)
    
    for l in LIST:
        try:
            I.pop(l)
        except: 
            pass
    I.write_file(inc_file)

def remove_vdw_SCAN():
    remove_tags_from_INCAR(LIST = ['BPARAM','LUSE_VDW'])

def Z2_kpoints(kpt=[],POSCAR='input/POSCAR',density=1000):
    #print (os.getcwd())
    S = Structure.from_file(POSCAR)
    K = Kpoints.automatic_density(S,density)
    return str(K)

def compute9(S, tol = 1.1,seed_index=0, write_poscar_from_cluster=False):
    S = SpacegroupAnalyzer(S,0.1).get_conventional_standard_structure()

    if len([e for e in S.composition if e.symbol in ['He', 'Ne', 'Ar', 'Kr', 'Xe']]) != 0:
        type = 'noble gas'
    else:
        S.make_supercell(2)
        Distance_matrix = S.distance_matrix
        np.fill_diagonal(Distance_matrix,100)
        radii = [Elem_radius[site.species_string] for site in S.sites]
        radiiT = np.array(radii)[np.newaxis].T
        radii_matrix = radii + radiiT*tol
        temp = Distance_matrix -radii_matrix
        binary_matrix = (temp < 0).astype(int)


        seed = set((np.where( binary_matrix[seed_index]==1 ))[0])
        cluster  = seed
        NEW = seed
        check = True
        while check:
            Ss = set()
            for n in NEW:
                Ss.update(set(np.where( binary_matrix[n]==1 )[0]))
            if Ss.issubset(cluster):
                check = False
            else:
                NEW = Ss - cluster
                cluster.update(Ss)
        if len(cluster) == 0:  # i.e. the cluster is a single atom.
            type = 'molecular'

        elif len(cluster) == len(S.sites): # i.e. all atoms are bonded.
            type = 'conventional'

        else:
            uniform = True
            cmp = Composition.from_dict(Counter([S[l].specie.name for l in list(cluster)]))
            if cmp.reduced_formula != S.composition.reduced_formula:
                    uniform = False
            if not uniform:
                type = 'heterogeneous'
            else:
                old_cluster_size = len(cluster)
                S.make_supercell(2)
                Distance_matrix = S.distance_matrix
                np.fill_diagonal(Distance_matrix,100)
                radii = [Elem_radius[site.species_string] for site in S.sites]
                radiiT = np.array(radii)[np.newaxis].T
                radii_matrix = radii + radiiT*tol
                temp = Distance_matrix-radii_matrix
                binary_matrix = (temp < 0).astype(int)

                seed = set((np.where(binary_matrix[seed_index]==1))[0])
                cluster  = seed
                NEW = seed
                check = True
                while check:
                    Ss = set()
                    for n in NEW:
                        Ss.update(set(np.where( binary_matrix[n]==1 )[0]))

                    if Ss.issubset(cluster):
                        check = False
                    else:
                        NEW = Ss - cluster
                        cluster.update(Ss)

               # print len(cluster)
                if len(cluster) != 4 * old_cluster_size:
                    type = 'molecular'

                else:
                    type = 'layered'
    if write_poscar_from_cluster:
        S.from_sites(cluster).to('POSCAR', 'POSCAR')
    return type
